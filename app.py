import os
import tempfile
import time
import yaml
import numpy as np
import pandas as pd
import streamlit as st

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter

from validator import (
    NotBlankValidator,
    TypeValidator,
    LengthValidator,
    RegexValidator,
    EmailValidator,
    ChoiceValidator,
    DateTimeValidator,
    ExcelDateValidator,
    CountryValidator,
    ConditionalValidator,
    URLValidator,
    NumericRangeValidator,
    DateRangeValidator
)


# ---------------------------------------------------------
# Page Configuration
# ---------------------------------------------------------

st.set_page_config(
    page_title="Excel Validation Engine",
    page_icon="✓",
    layout="wide"
)


# ---------------------------------------------------------
# Validator Mapping
# ---------------------------------------------------------

CLASSMAP = {
    "NotBlank": NotBlankValidator.NotBlankValidator,
    "Type": TypeValidator.TypeValidator,
    "Length": LengthValidator.LengthValidator,
    "Regex": RegexValidator.RegexValidator,
    "Email": EmailValidator.EmailValidator,
    "Choice": ChoiceValidator.ChoiceValidator,
    "Date": DateTimeValidator.DateTimeValidator,
    "ExcelDate": ExcelDateValidator.ExcelDateValidator,
    "Country": CountryValidator.CountryValidator,
    "Conditional": ConditionalValidator.ConditionalValidator,
    "URL": URLValidator.URLValidator,
    "NumericRange": NumericRangeValidator.NumericRangeValidator,
    "DateRange": DateRangeValidator.DateRangeValidator
}


# ---------------------------------------------------------
# Settings
# ---------------------------------------------------------

def set_settings(config_data):

    settings = {}

    excludes = []

    config = yaml.safe_load(config_data)

    if config is None:
        raise ValueError("Empty YAML configuration.")

    if "validators" not in config:
        raise ValueError("Missing 'validators' section.")

    if "columns" not in config["validators"]:
        raise ValueError("Missing 'validators.columns' section.")

    settings["validators"] = config["validators"]["columns"]

    if "default" in config["validators"]:
        settings["defaultValidator"] = config["validators"]["default"][0]
    else:
        settings["defaultValidator"] = None

    if "excludes" in config:

        for column in config["excludes"]:
            excludes.append(column_index_from_string(column))

        settings["excludes"] = excludes

    else:
        settings["excludes"] = []

    if "range" in config:
        settings["range"] = (
            config["range"][0]
            + "1:"
            + config["range"][1]
        )
    else:
        settings["range"] = None

    if "header" in config:
        settings["header"] = config["header"]
    else:
        settings["header"] = True

    return settings


# ---------------------------------------------------------
# Empty Row
# ---------------------------------------------------------

def is_empty(row):

    for cell in row:
        if cell.value is not None and cell.value != "":
            return False

    return True


# ---------------------------------------------------------
# Validation
# ---------------------------------------------------------

def is_valid(validation_type, value, coordinate, errors, value2=None):

    name = list(validation_type.keys())[0]
    data = list(validation_type.values())[0]

    if name not in CLASSMAP:
        errors.append(
            (
                coordinate,
                ["Unknown validator: " + name]
            )
        )

        return False

    validator = CLASSMAP[name](data)

    if name != "Conditional":
        result = validator.validate(value)
    else:
        result = validator.validate(value, value2)

    if result is False:

        errors.append(
            (
                coordinate,
                [validator.getMessage()]
            )
        )

        return False

    return True


# ---------------------------------------------------------
# Validate Excel
# ---------------------------------------------------------

def validate_excel(settings, excel_path, sheet_name, progress_bar):

    errors = []

    wb = load_workbook(
        excel_path,
        keep_vba=True,
        data_only=False
    )

    ws = wb[sheet_name]

    max_rows = ws.max_row

    validation_range = settings.get("range")

    if validation_range is not None:
        validation_range = validation_range + str(max_rows)

    processed_rows = 0

    rows = ws.iter_rows(validation_range)

    for row in rows:

        processed_rows += 1

        progress = min(
            processed_rows / max_rows,
            1.0
        )

        progress_bar.progress(
            progress,
            text=f"Validating row {processed_rows} of {max_rows}"
        )

        if is_empty(row):
            continue

        column_counter = 0

        for cell in row:

            column_counter += 1

            value = cell.value

            # Header detection
            if settings["header"] is not True:

                if value == settings["header"]:
                    settings["header"] = True

                break

            # Excluded column
            if (
                hasattr(cell, "column")
                and cell.column in settings["excludes"]
            ):
                continue

            column = get_column_letter(column_counter)

            coordinate = "{}{}".format(
                column,
                cell.row
            )

            # Specific column validators
            if column in settings["validators"]:

                validators = settings["validators"][column]

                for validation_type in validators:

                    name = list(validation_type.keys())[0]

                    if name != "Conditional":

                        result = is_valid(
                            validation_type,
                            value,
                            coordinate,
                            errors
                        )

                    else:

                        field_b = list(
                            validation_type.values()
                        )[0]["fieldB"]

                        value2 = ws[
                            field_b + str(cell.row)
                        ].value

                        result = is_valid(
                            validation_type,
                            value,
                            coordinate,
                            errors,
                            value2
                        )

                    if not result:
                        break

            # Default validator
            elif settings["defaultValidator"] is not None:

                is_valid(
                    settings["defaultValidator"],
                    value,
                    coordinate,
                    errors
                )

    progress_bar.progress(
        1.0,
        text="Validation completed"
    )

    wb.close()

    return errors


# ---------------------------------------------------------
# Mark Errors
# ---------------------------------------------------------

def create_error_file(
    errors,
    original_file,
    sheet_name,
    print_errors
):

    file_size = os.path.getsize(original_file)

    # 10 MB limit
    if file_size > 10485760:

        raise ValueError(
            "File is larger than 10 MB. "
            "Annotated Excel generation is disabled."
        )

    extension = os.path.splitext(
        original_file
    )[1]

    keep_vba = extension.lower() == ".xlsm"

    wb = load_workbook(
        original_file,
        keep_vba=keep_vba,
        data_only=False
    )

    ws = wb[sheet_name]

    red_fill = PatternFill(
        start_color="FFFF0000",
        end_color="FFFF0000",
        fill_type="solid"
    )

    for coordinate, violations in errors:

        cell = ws[coordinate]

        cell.fill = red_fill

        if print_errors:

            cell.value = ", ".join(
                violations
            )

    output_file = tempfile.NamedTemporaryFile(
        delete=False,
        suffix=extension
    )

    output_file.close()

    wb.save(output_file.name)

    wb.close()

    return output_file.name


# ---------------------------------------------------------
# UI
# ---------------------------------------------------------

st.title("Excel Validation Engine")

st.caption(
    "Validate Excel files using configurable YAML validation rules."
)

st.divider()


# ---------------------------------------------------------
# Sidebar
# ---------------------------------------------------------

with st.sidebar:

    st.header("Validation Settings")

    st.write(
        "Upload an Excel file and a YAML configuration."
    )

    st.info(
        "Supported validators: "
        "NotBlank, Type, Length, Regex, Email, "
        "Choice, Date, ExcelDate, Country, Conditional, "
        "URL, NumericRange and DateRange."
    )


# ---------------------------------------------------------
# File Upload
# ---------------------------------------------------------

col1, col2 = st.columns(2)

with col1:

    excel_file = st.file_uploader(
        "Upload Excel File",
        type=["xlsx", "xlsm"]
    )

with col2:

    yaml_file = st.file_uploader(
        "Upload YAML Configuration",
        type=["yaml", "yml"]
    )


# ---------------------------------------------------------
# Main
# ---------------------------------------------------------

if excel_file and yaml_file:

    st.divider()

    # Save uploaded Excel temporarily

    excel_temp = tempfile.NamedTemporaryFile(
        delete=False,
        suffix=os.path.splitext(
            excel_file.name
        )[1]
    )

    excel_temp.write(
        excel_file.getbuffer()
    )

    excel_temp.close()

    # Read workbook sheets

    try:

        wb = load_workbook(
            excel_temp.name,
            read_only=True,
            keep_vba=True
        )

        sheet_names = wb.sheetnames

        wb.close()

    except Exception as e:

        st.error(
            "Unable to read Excel file: {}".format(e)
        )

        st.stop()


    # -----------------------------------------------------
    # File Information
    # -----------------------------------------------------

    st.subheader("File Information")

    info1, info2, info3 = st.columns(3)

    with info1:
        st.metric(
            "File",
            excel_file.name
        )

    with info2:
        st.metric(
            "Size",
            "{:.2f} MB".format(
                len(excel_file.getbuffer()) / 1024 / 1024
            )
        )

    with info3:
        st.metric(
            "Sheets",
            len(sheet_names)
        )


    # -----------------------------------------------------
    # Sheet Selection
    # -----------------------------------------------------

    sheet_name = st.selectbox(
        "Select Worksheet",
        sheet_names
    )


    # -----------------------------------------------------
    # YAML Preview
    # -----------------------------------------------------

    with st.expander(
        "View YAML Configuration"
    ):

        yaml_content = yaml_file.getvalue().decode(
            "utf-8"
        )

        st.code(
            yaml_content,
            language="yaml"
        )


    # -----------------------------------------------------
    # Validate Button
    # -----------------------------------------------------

    st.divider()

    validate_button = st.button(
        "Validate Excel File",
        type="primary",
        use_container_width=True
    )


    if validate_button:

        try:

            # Parse YAML

            yaml_content = yaml_file.getvalue().decode(
                "utf-8"
            )

            settings = set_settings(
                yaml_content
            )

            st.success(
                "Validation configuration loaded successfully."
            )

            # Progress

            progress_bar = st.progress(
                0,
                text="Starting validation..."
            )

            start_time = time.time()

            errors = validate_excel(
                settings,
                excel_temp.name,
                sheet_name,
                progress_bar
            )

            execution_time = time.time() - start_time

            progress_bar.progress(
                1.0,
                text="Validation completed"
            )


            # -------------------------------------------------
            # Results
            # -------------------------------------------------

            st.divider()

            st.subheader("Validation Results")

            total_errors = len(errors)

            if total_errors == 0:

                st.success(
                    "Validation passed. No errors found."
                )

                col1, col2 = st.columns(2)

                with col1:
                    st.metric(
                        "Errors",
                        0
                    )

                with col2:
                    st.metric(
                        "Execution Time",
                        "{:.2f}s".format(
                            execution_time
                        )
                    )

            else:

                st.error(
                    "Validation failed. {} error(s) found.".format(
                        total_errors
                    )
                )

                col1, col2, col3 = st.columns(3)

                with col1:
                    st.metric(
                        "Errors",
                        total_errors
                    )

                with col2:
                    st.metric(
                        "Worksheet",
                        sheet_name
                    )

                with col3:
                    st.metric(
                        "Execution Time",
                        "{:.2f}s".format(
                            execution_time
                        )
                    )


                # ---------------------------------------------
                # Error Table
                # ---------------------------------------------

                st.subheader("Validation Errors")

                error_rows = []

                for coordinate, violations in errors:

                    error_rows.append(
                        {
                            "Cell": coordinate,
                            "Error": ", ".join(
                                violations
                            )
                        }
                    )

                error_frame = pd.DataFrame(error_rows)
                error_frame["Column"] = error_frame["Cell"].str.extract(
                    r"([A-Z]+)",
                    expand=False
                )
                errors_per_column = error_frame["Column"].value_counts()
                error_distribution = errors_per_column.to_numpy(dtype=float)

                st.dataframe(
                    error_frame,
                    use_container_width=True,
                    hide_index=True
                )

                summary_col1, summary_col2, summary_col3 = st.columns(3)

                with summary_col1:
                    st.metric(
                        "Affected Columns",
                        int(error_frame["Column"].nunique())
                    )

                with summary_col2:
                    st.metric(
                        "Avg Errors / Column",
                        "{:.1f}".format(np.mean(error_distribution))
                    )

                with summary_col3:
                    st.metric(
                        "Median Errors / Column",
                        "{:.1f}".format(np.median(error_distribution))
                    )

                st.download_button(
                    "Download Errors CSV",
                    data=error_frame.to_csv(index=False).encode("utf-8"),
                    file_name="validation_errors.csv",
                    mime="text/csv"
                )


                # ---------------------------------------------
                # Generate Error Excel
                # ---------------------------------------------

                st.subheader(
                    "Generate Error Report"
                )

                print_errors = st.checkbox(
                    "Write validation messages inside failed cells",
                    value=True
                )

                if st.button(
                    "Generate Annotated Excel",
                    use_container_width=True
                ):

                    try:

                        with st.spinner(
                            "Generating annotated Excel file..."
                        ):

                            error_file = create_error_file(
                                errors,
                                excel_temp.name,
                                sheet_name,
                                print_errors
                            )

                        with open(
                            error_file,
                            "rb"
                        ) as file:

                            st.download_button(
                                label="Download Error Excel",
                                data=file,
                                file_name=(
                                    "validation_errors_"
                                    + excel_file.name
                                ),
                                mime=(
                                    "application/vnd.openxmlformats-officedocument"
                                    ".spreadsheetml.sheet"
                                ),
                                use_container_width=True
                            )

                        st.success(
                            "Error Excel file generated successfully."
                        )

                    except Exception as e:

                        st.error(
                            "Could not generate error file: {}".format(e)
                        )

        except Exception as e:

            st.error(
                "Validation failed: {}".format(e)
            )


    # -----------------------------------------------------
    # Cleanup
    # -----------------------------------------------------

    if os.path.exists(excel_temp.name):

        try:
            os.unlink(excel_temp.name)
        except:
            pass

else:

    st.info(
        "Upload an Excel file and YAML configuration to begin."
    )